import os
import pandas as pd
import sys
import re
import shutil
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font

def iperfserverfiling(filename, commandline):
    now = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')

    fread = open(filename + '.txt', 'r', encoding='utf-8')
    fwrite = open(filename + '_' + 'temp.txt', 'w', encoding='utf-8')

    while(1):
        v = fread.readline()
        if v.__contains__("Jitter"):
            fwrite.write("interval,sec,transfer,bytes,bitrate,bits/sec,jitter,remarks\n")
        
        elif v == "\n":
            break
        elif v == "":
            break


    fwrite.close()
    fread.close()

    df = pd.read_csv(filename + '_' + 'temp.txt', encoding='utf-8', header=0)
    df.index += 1
    df.to_excel(filename + '.xlsx', index = True)

    wb = load_workbook(filename + '.xlsx', data_only=True)
    ws = wb['Sheet1']

    border_thin = Side(border_style='thin')
    all_border_thin = Border(left = border_thin, right = border_thin, top = border_thin, bottom = border_thin)
    lines = len(df)

    for i in range(lines + 2, lines + 5):
        ws.merge_cells(start_row = i, start_column = 2, end_row = i, end_column = 5)
        ws.cell(row = i, column = 1).border = all_border_thin
        ws.cell(row = i, column = 1).font = Font(bold = True)

    ws.cell(row = lines + 2, column = 1).value = "CMD"
    ws.cell(row = lines + 3, column = 1).value = "일시"
    ws.cell(row = lines + 4, column = 1).value = "비고"
    ws.cell(row = lines + 2, column = 2).value = commandline
    ws.cell(row = lines + 3, column = 2).value = date + ' ' + time

    for i in ['A', 'B', 'C', 'D', 'E']:
        for cell in range(len(ws[i])):
            ws[i + str(cell + 1)].alignment = Alignment(horizontal = 'center')
            if type(ws[i + str(cell + 1)].value) == str:
                try:
                    ws[i + str(cell + 1)].value = int(ws[i + str(cell + 1)].value)
                except:
                    try:
                        ws[i + str(cell + 1)].value = float(ws[i + str(cell + 1)].value)
                    except:
                        continue

    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            ws.cell(row = i, column = j).border = all_border_thin

    ws['D1'].value = ''
    ws['E1'].value = ''

    wb.save('results/' + filename + '.xlsx')
    os.makedirs('results/txt', exist_ok=True)
    shutil.move(filename + '.txt', 'results/txt')
    os.remove(filename + '_' + 'temp.txt')
    os.remove(filename + '.xlsx')

def arrange_time(line):
    try:
        result = re.findall(r'\d+(?:\.\d+)*', line)[-1]
    except:
        result = "no data"

    return result

def arrange(line):
    try:
        result = re.findall(r'\d+\.?\d*', line)
    except:
        result = ["0", "0", "0", "0"]

    return result

if __name__ == "__main__":
    iperfserverfiling("iperf3_server", commandline="iperf3 -s")