import os
import pandas as pd
import sys
import re
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font

def pingfiling(filename, commandline):
    filename_list = filename.split('_')
    date = filename_list[1]
    time = filename_list[2]

    fread = open(filename + '.txt', 'r', encoding='utf-8')
    fwrite = open(filename + '_' + 'temp.txt', 'w', encoding='utf-8')

    fwrite.write("time(ms),-,--\n")
    fread.seek(0)
    fread.seek(0)

    while(1):
        v = fread.readline()

        if v.__contains__("만료") or v.__contains__("timed out"):
            fwrite.write("timeout,,\n")
        elif v.__contains__("연결할 수 없습니다") or v.__contains__("failed") or v.__contains__("unreachable"):
            fwrite.write("unreachable,,\n")
        elif v.__contains__("시간") or v.__contains__("Reply from"):
            v = arrange_time(v)
            fwrite.write(v + ",,\n")
        elif v == "\n":
            break
        elif v == "":
            sys.exit()
        
    while(1):
        v = fread.readline()

        if v.__contains__("보냄") or v.__contains__("Sent"):
            v = arrange(v)
            fwrite.write("send,receive,loss\n")
            fwrite.write(','.join(v[:3]) + '\n')
        
        elif v.__contains__("최소") or v.__contains__("Minimum"):
            v = arrange(v)
            fwrite.write("min,max,avg\n")
            fwrite.write(','.join(v[:3]) + '\n')

        elif v == "\n":
            break

        elif v == "":
            break


    fwrite.close()
    fread.close()

    df = pd.read_csv(filename + '_' + 'temp.txt', encoding='utf-8', header=0)
    df.index += 1
    df.to_excel(filename + '.xlsx', index = True)

    wb = load_workbook(filename + '.xlsx', data_only=True)
    ws = wb['Sheet1']

    border_thin = Side(border_style='thin')
    all_border_thin = Border(left = border_thin, right = border_thin, top = border_thin, bottom = border_thin)
    lines = len(df)

    for i in range(lines + 2, lines + 5):
        ws.merge_cells(start_row = i, start_column = 2, end_row = i, end_column = 4)
        ws.cell(row = i, column = 1).border = all_border_thin
        ws.cell(row = i, column = 1).font = Font(bold = True)

    ws.cell(row = lines + 2, column = 1).value = "CMD"
    ws.cell(row = lines + 3, column = 1).value = "일시"
    ws.cell(row = lines + 4, column = 1).value = "비고"
    ws.cell(row = lines + 2, column = 2).value = commandline
    ws.cell(row = lines + 3, column = 2).value = date + ' ' + time

    for i in ['A', 'B', 'C', 'D']:
        for cell in range(len(ws[i])):
            ws[i + str(cell + 1)].alignment = Alignment(horizontal = 'center')
            if type(ws[i + str(cell + 1)].value) == str:
                try:
                    ws[i + str(cell + 1)].value = int(ws[i + str(cell + 1)].value)
                except:
                    continue

    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            ws.cell(row = i, column = j).border = all_border_thin

    ws['C1'].value = ''
    ws['D1'].value = ''

    wb.save('results/' + filename + '.xlsx')
    os.makedirs('results/txt', exist_ok=True)
    shutil.move(filename + '.txt', 'results/txt')
    os.remove(filename + '_' + 'temp.txt')
    os.remove(filename + '.xlsx')

def arrange(line):
    try:
        result = re.findall(r'\d+', line)
    
    except:
        result = ['0' for _ in range(4)]

    return result

def arrange_time(line):
    try:
        line_tmp = line.split("시간")[1] # change "시간" to "time" if you use not KR, but EN
        result = re.findall(r'\d+', line_tmp)[0]

    except:
        result = "no data"

    return result

if __name__ == "__main__":
    pingfiling("test", commandline="ping google.com -n 10")