import pandas as pd
import os
import sys
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font

def iperfiling(filename, commandline) :
    filename_list = filename.split('_')
    date = filename_list[1]
    time = filename_list[2]

    fread = open(filename + '.txt', 'r', encoding='utf-8')
    fwrite = open(filename + '_' + 'temp.txt', 'w', encoding='utf-8')

    fwrite.write("interval,sec,transfer,bytes,bitrate,bits/sec,remarks\n")

    while(1):
        v = fread.readline()

        if v.__contains__("sec"):
            v = arrange(v)
            fwrite.write(','.join(v[1:7]) + ',\n')
        elif v.__contains__("- - -") or v.__contains__("---"):
            fwrite.write('-,-,-,-,-,-,-\n')
            break        
        elif v == "\n":
            break
        elif v == "":
            sys.exit()

    while(1):
        v = fread.readline()

        if v.__contains__("sender") or v.__contains__("receiver"):
            v = arrange(v)
            fwrite.write(','.join(v[1:8]) + '\n')
        elif v == "\n":
            break
        elif v == "":
            break
        
    fwrite.close()
    fread.close()

    df = pd.read_csv(filename + '_temp.txt', sep = ',', encoding = 'utf-8', header = 0)
    df.index += 1
    df.to_excel(filename + '.xlsx', index = True)

    wb = load_workbook(filename + '.xlsx', data_only=True)
    ws = wb['Sheet1']

    border_thin = Side(border_style='thin')
    all_border_thin = Border(left = border_thin, right = border_thin, top = border_thin, bottom = border_thin)
    lines = df.index[-1]

    for i in range(lines + 2, lines + 5):
        ws.merge_cells(start_row = i, start_column = 2, end_row = i, end_column = 8)
        ws.cell(row = i, column = 1).border = all_border_thin
        ws.cell(row = i, column = 1).font = Font(bold = True)

    ws.cell(row = lines + 2, column = 1).value = "CMD"
    ws.cell(row = lines + 3, column = 1).value = "일시"
    ws.cell(row = lines + 4, column = 1).value = "비고"
    ws.cell(row = lines + 2, column = 2).value = commandline
    ws.cell(row = lines + 3, column = 2).value = date + ' ' + time

    for i in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        for cell in range(len(ws[i])):
            ws[i + str(cell + 1)].alignment = Alignment(horizontal = 'center')
            if type(ws[i + str(cell + 1)].value) == str:
                try:
                    ws[i + str(cell + 1)].value = int(ws[i + str(cell + 1)].value)
                except:
                    try:
                        ws[i + str(cell + 1)].value = float(ws[i + str(cell + 1)].value)
                    except:
                        continue
                    
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            ws.cell(row = i, column = j).border = all_border_thin
    
    ws.merge_cells(start_row = lines - 1, start_column = 2, end_row = lines - 1, end_column = 8)

    wb.save('results/' + filename + '.xlsx')
    os.makedirs('results/txt', exist_ok=True)
    shutil.move(filename + '.txt', 'results/txt')
    os.remove(filename + '_' + 'temp.txt')
    os.remove(filename + '.xlsx')

def iperfiling_udp(filename, commandline) :
    filename_list = filename.split('_')
    date = filename_list[1]
    time = filename_list[2]

    fread = open(filename + '.txt', 'r', encoding='utf-8')
    fwrite = open(filename + '_' + 'temp.txt', 'w', encoding='utf-8')

    fwrite.write("interval,sec,transfer,bytes,bitrate,bits/sec,total,remarks\n")

    while(1):
        v = fread.readline()

        if v.__contains__("sec"):
            v = arrange(v)
            fwrite.write(','.join(v[1:8]) + ',\n')
        elif v.__contains__("- - -") or v.__contains__("---"):
            fwrite.write('-,-,-,-,-,-,-,-\n')
            fwrite.write("interval,sec,transfer,bytes,bitrate,bits/sec,jitter,remarks\n")
            break        
        elif v == "\n":
            break
        elif v == "":
            sys.exit()

    while(1):
        v = fread.readline()

        if v.__contains__("sender"):
            v = arrange(v)
            fwrite.write(','.join(v[1:8]) + ',sender\n')
        elif v.__contains__("receiver"):
            v = arrange(v)
            fwrite.write(','.join(v[1:8]) + ',receiver\n')
        elif v == "\n":
            break
        elif v == "":
            break
            
    fwrite.close()
    fread.close()

    df = pd.read_csv(filename+ '_temp.txt', sep=',', encoding='utf-8', header=0)
    df.index += 1
    df.to_excel(filename + '.xlsx', index = True)

    wb = load_workbook(filename + '.xlsx', data_only=True)
    ws = wb['Sheet1']

    border_thin = Side(border_style='thin')
    all_border_thin = Border(left = border_thin, right = border_thin, top = border_thin, bottom = border_thin)
    lines = df.index[-1]

    for i in range(lines + 2, lines + 5):
        ws.merge_cells(start_row = i, start_column = 2, end_row = i, end_column = 9)
        ws.cell(row = i, column = 1).border = all_border_thin
        ws.cell(row = i, column = 1).font = Font(bold = True)

    ws.cell(row = lines + 2, column = 1).value = "CMD"
    ws.cell(row = lines + 3, column = 1).value = "일시"
    ws.cell(row = lines + 4, column = 1).value = "비고"
    ws.cell(row = lines + 2, column = 2).value = commandline
    ws.cell(row = lines + 3, column = 2).value = date + ' ' + time

    for i in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        for cell in range(len(ws[i])):
            ws[i + str(cell + 1)].alignment = Alignment(horizontal = 'center')
            if type(ws[i + str(cell + 1)].value) == str:
                try:
                    ws[i + str(cell + 1)].value = int(ws[i + str(cell + 1)].value)
                except:
                    try:
                        ws[i + str(cell + 1)].value = float(ws[i + str(cell + 1)].value)
                    except:
                        continue

    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            ws.cell(row = i, column = j).border = all_border_thin
    
    ws.merge_cells(start_row = lines - 2, start_column = 2, end_row = lines - 2, end_column = 9)

    wb.save('results/' + filename + '.xlsx')
    os.makedirs('results/txt', exist_ok=True)
    shutil.move(filename + '.txt', 'results/txt')
    os.remove(filename + '_' + 'temp.txt')
    os.remove(filename + '.xlsx')

def arrange(line):
    try : 
        result = line.replace('[', '').replace(']', '').split()
    
    except:
        result = ['NaN' for _ in range(10)]
    
    return result

if __name__ == "__main__":
    iperfiling("test", commandline="iperf -c localhost -t 5")
